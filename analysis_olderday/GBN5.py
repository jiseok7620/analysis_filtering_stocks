import openpyxl
import pandas as pd


class GBN5_cls:
    def Exe_GBN5(self,name,dd):
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('F:/JusikData/analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']


        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "F:/JusikData/oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        print(data1.columns)
        
        
        
        ##---------------------------------------------------------------------##
        # 변수에 데이터 저장
        # 해당일자 시가총액의 인덱스구하기
        시가총액_id = data1.loc[data1['일자'] == dd]['시가총액'].index[0]
        
        arr_dm = [0,1,2,3,4,5,6,7]
        
        for i in arr_dm:
            시가총액 = data1.iloc[시가총액_id-i]['시가총액']
            print(시가총액)
        
            ##---------------------------------------------------------------------##
            # 좌표로 값 출력 = sheet.cell(행, 열, 값)
            sheet.cell(329, 3 + i, 시가총액)
        
        
        
        ##---------------------------------------------------------------------##
        # 저장
        wb.save('F:/JusikData/analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
#conn = GBN5_cls()
#conn.Exe_GBN5('삼성전자',20200504)