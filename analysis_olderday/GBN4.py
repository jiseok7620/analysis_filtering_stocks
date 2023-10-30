import openpyxl
import pandas as pd
import traceback
import datetime
import numpy as np
from Invest.analysis_olderday.except_hol import except_hol_cls

class GBN4_cls:
    # name = 종목명, dd = 일자
    def Exe_GBN4(self,name,dd):
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('F:/JusikData/analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']
        
        print('여기는 GBN4')

        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "F:/JusikData/oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        #print(data1.columns)
        
        ## data2 = 재무 정보 데이터 ##
        path2_1 = "F:/JusikData/report_csv/report_make/재무정보_1분기.csv"
        path2_2 = "F:/JusikData/report_csv/report_make/재무정보_2분기.csv"
        path2_3 = "F:/JusikData/report_csv/report_make/재무정보_3분기.csv"
        path2_4 = "F:/JusikData/report_csv/report_make/재무정보_4분기.csv"
        data2_1 = pd.read_csv(path2_1, encoding='cp949')
        data2_2 = pd.read_csv(path2_2, encoding='cp949')
        data2_3 = pd.read_csv(path2_3, encoding='cp949')
        data2_4 = pd.read_csv(path2_4, encoding='cp949')
        #print(data2_1.columns)
        
        ## data4 = 업종 분류 데이터 ##
        path4 = "F:/JusikData/oneday_csv/data_sectors/"+str(dd)[0:4]+'.csv'
        data4 = pd.read_csv(path4, encoding='cp949')
        data4 = data4.drop(['종가','대비', '등락률', '시가총액'], axis=1)
        #print(data4.columns)
        
        ## 데이터 만들기
        # 현재가, 당기순이익, 상장주식수, 순자산, 자기자본, 자산을 data4에 추가해주기 위해서
        data2_1 = data2_1.drop(['종목코드','업종명'], axis=1)
        data2_2 = data2_2.drop(['종목코드','업종명'], axis=1)
        data2_3 = data2_3.drop(['종목코드','업종명'], axis=1)
        data2_4 = data2_4.drop(['종목코드','업종명'], axis=1)
        
        data5_1 = pd.merge(data4,data2_1, how='outer', on=['종목명'])
        data5_2 = pd.merge(data4,data2_2, how='outer', on=['종목명'])        
        data5_3 = pd.merge(data4,data2_3, how='outer', on=['종목명'])
        data5_4 = pd.merge(data4,data2_4, how='outer', on=['종목명'])
        #print(data5_1.columns)
        
        
        
        ##---------------------------------------------------------------------##
        ## 변수에 데이터 저장
        # 당해 구하기(과거 데이터 분석 시)
        dd_str = str(dd)[0:4]
        dd_int = int(dd_str)
        
        # dd의 년, 월, 일 나누기
        dd_y = int(str(dd)[0:4])
        dd_m = int(str(dd)[4:6])
        dd_d = int(str(dd)[6:8])
        
        # 년도 배열 만들기
        arr_yago = [dd_y,dd_y-1,dd_y-2,dd_y-3,dd_y-4,dd_y-5]
        
        # 행, 열 더하기
        col_num = 0
        
        # 년도, 분기 반복문 돌리기        
        for j in arr_yago:
            # 2015년의 데이터는 없으므로 2015년 이후부터만 실행
            if j > 2015 and j != dd_y :
                ## 일자 데이터 만들기
                # 일자 구하기
                dd_full = str(j)[0:4] + str(dd)[4:8]
                
                # 요일이 토요일이나 일요일또는 공휴일 일때 날짜 바꿔주기
                dd_full = except_hol_cls.exe_except(self, dd_full)
                        
                # 해당 종목의 업종명 구하기
                try:
                    up_name = data4.loc[data4['종목명'] == name]['업종명'].values[0]
                except:
                    up_name = 'non data'
                    
                # 업종명이 있다면 지표를 계산 하자
                if up_name != 'non data' :
                    
                    # 해당 업종을 가진 종목들을 구해서 list로 만들기
                    업종_종목명 = data5_4.loc[data5_4['업종명'] == up_name]['종목명'].tolist()
                    
                    ##---------------------------------------------------------------------##
                    # GB4-1. PER 구하기
                    arr_new업종_1 = []
                    arr_new업종_1_2 = []                
                    arr_업종현재가 = []
                    arr_업종당기순이익 = []
                    arr_업종상장주식수 = []
                    for i in 업종_종목명:
                        # 해당 업종의 당기순이익 구하기
                        try:
                            당기순이익 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_당기순이익_4'].values[0]
                            if np.isnan(당기순이익) :
                                당기순이익 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_당기순이익(손익)_4'].values[0]  
                                if np.isnan(당기순이익) :
                                    당기순이익 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_당기순이익(2)_4'].values[0]                    
                        except:
                            당기순이익 = 'non data'
                        
                        arr_업종당기순이익.append(당기순이익)
                        
                        if 당기순이익 != 'non data' :
                            if np.isnan(당기순이익) == False :
                                arr_new업종_1.append(i)
                    
                    for i in arr_new업종_1 :
                        # 해당 업종의 현재가 구하기
                        # 해당 업종의 상장주식수 구하기
                        path_업종 = "F:/JusikData/oneday_csv/onedaydata/"+i+'/'+i+'.csv'
                        data_업종 = pd.read_csv(path_업종, encoding='cp949')
                        try:
                            arr_업종현재가.append(data_업종.loc[data_업종['일자']==int(dd_full)]['종가'].values[0])
                            arr_업종상장주식수.append(data_업종.loc[data_업종['일자']==int(dd_full)]['상장주식수'].values[0])
                        except:
                            pass
                    
                    # 업종 현재가 평균 구하기
                    업종현재가 = np.mean(arr_업종현재가)
                    
                    #업종 당기순이익 평균 구하기
                    arr_업종당기순이익 = [x for x in arr_업종당기순이익 if x != 'nan data'] # nan data 값 제거
                    arr_업종당기순이익 = [x for x in arr_업종당기순이익 if np.isnan(x) == False] # nan 값 제거
                    업종당기순이익 = np.mean(arr_업종당기순이익)
                    
                    # 업종 상장주식수 평균 구하기
                    업종상장주식수 = np.mean(arr_업종상장주식수)
                    
                    # 확인 작업
                    #print(업종현재가,업종당기순이익,업종상장주식수)
                    #print(len(arr_업종현재가),len(arr_업종당기순이익),len(arr_업종상장주식수))
                    
                    # 업종_PER 구하기
                    업종_PER = 업종현재가 / (업종당기순이익 / 업종상장주식수)
                    print('업종현재가: ',arr_업종현재가)
                    print('업종당기순이익: ',arr_업종당기순이익)
                    print('업종상장주식수: ',arr_업종상장주식수)
                    print('크기비교(세개가같아야함): ', len(arr_업종현재가), len(arr_업종당기순이익), len(arr_업종상장주식수))
                    print('업종PER : ',업종_PER)
                    
    
                    ##---------------------------------------------------------------------##
                    # GB4-2. PBR 구하기
                    arr_new업종_2 = []                
                    arr_업종현재가 = []
                    arr_업종자본총계 = []
                    arr_업종상장주식수 = []
                    for i in 업종_종목명:
                        # 해당 업종의 자본총계 구하기
                        try:
                            자본총계 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_자본총계_4'].values[0]
                            if np.isnan(자본총계) :
                                자본총계 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_자본총계(2)_4'].values[0]
                        except:
                            자본총계 = 'non data'
                        
                        arr_업종자본총계.append(자본총계)
                        
                        if 자본총계 != 'non data' :
                            if np.isnan(자본총계) == False:
                                arr_new업종_2.append(i)
                    
                    for i in arr_new업종_2 :
                        # 해당 업종의 현재가 구하기
                        # 해당 업종의 상장주식수 구하기
                        path_업종 = "F:/JusikData/oneday_csv/onedaydata/"+i+'/'+i+'.csv'
                        data_업종 = pd.read_csv(path_업종, encoding='cp949')
                        try:
                            arr_업종현재가.append(data_업종.loc[data_업종['일자']==int(dd_full)]['종가'].values[0])
                            arr_업종상장주식수.append(data_업종.loc[data_업종['일자']==int(dd_full)]['상장주식수'].values[0])
                        except:
                            pass
                            #logging.error(traceback.format_exc())
                            #dd일 이전에 상폐되었거나 거래정지면 에러가 발생
                    
                    # 업종 현재가 평균 구하기
                    업종현재가 = np.mean(arr_업종현재가)
                    
                    #업종 자본총계 평균 구하기
                    arr_업종자본총계 = [x for x in arr_업종자본총계 if x != 'nan data'] # nan data 값 제거
                    arr_업종자본총계 = [x for x in arr_업종자본총계 if np.isnan(x) == False] # nan 값 제거
                    업종자본총계 =np.mean(arr_업종자본총계)
                    
                    # 업종 상장주식수 평균 구하기
                    업종상장주식수 = np.mean(arr_업종상장주식수)
                    
                    # 확인 작업
                    #print(업종현재가,업종자본총계,업종상장주식수)
                    #print(len(arr_업종현재가),len(arr_업종자본총계),len(arr_업종상장주식수))
                    
                    # 업종 PBR 구하기
                    업종_PBR = 업종현재가 / (업종자본총계 / 업종상장주식수)
                    print('업종현재가: ',arr_업종현재가)
                    print('업종자본총계: ',arr_업종자본총계)
                    print('업종상장주식수: ',arr_업종상장주식수)
                    print('크기비교(세개가같아야함): ', len(arr_업종현재가), len(arr_업종자본총계), len(arr_업종상장주식수))
                    print('업종_PBR : ', 업종_PBR)
                    
                    
                    ##---------------------------------------------------------------------##
                    # GB4-3. ROE 구하기
                    arr_업종자본총계2 = []
                    arr_new업종_3 = []
                    arr_업종당기순이익2 = []
                    for i in arr_new업종_1 :
                        # 해당 업종의 자본총계 구하기
                        try:
                            자본총계2 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_자본총계_4'].values[0]
                            if np.isnan(자본총계2) :
                                자본총계2 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_자본총계(2)_4'].values[0]
                        except:
                            자본총계2 = 'non data'
                        
                        arr_업종자본총계2.append(자본총계2)
                        
                        if 자본총계2 != 'non data' :
                            if np.isnan(자본총계2) == False :
                                arr_new업종_3.append(i)
                    
                    for i in arr_new업종_3 :
                        # 해당 업종의 당기순이익 구하기
                        try:
                            당기순이익2 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_당기순이익_4'].values[0]
                            if np.isnan(당기순이익2) :
                                당기순이익2 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_당기순이익(손익)_4'].values[0]  
                                if np.isnan(당기순이익2) :
                                    당기순이익2 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_당기순이익(2)_4'].values[0] 
                        except:
                            pass
                            
                        arr_업종당기순이익2.append(당기순이익2)
                            
                    #업종 자본총계 평균 구하기
                    arr_업종자본총계2 = [x for x in arr_업종자본총계2 if x != 'nan data'] # nan data 값 제거
                    arr_업종자본총계2 = [x for x in arr_업종자본총계2 if np.isnan(x) == False] # nan 값 제거
                    업종자본총계2 = np.mean(arr_업종자본총계2)
                    
                    #업종 당기순이익 평균 구하기
                    업종당기순이익2 = np.mean(arr_업종당기순이익2)
                    
                    # 업종 ROE 구하기
                    업종_ROE = 업종당기순이익2 / 업종자본총계2 * 100
                    print('업종당기순이익2: ',arr_업종당기순이익2)
                    print('업종자본총계2: ',arr_업종자본총계2)
                    print('크기비교(두개가같아야함): ', len(arr_업종당기순이익2), len(arr_업종자본총계2))
                    print('업종_ROE : ', 업종_ROE)
                    
                    
                    ##---------------------------------------------------------------------##
                    # GB4-4. ROA 구하기
                    arr_업종자산총계 = []
                    arr_new업종_4 = []
                    arr_업종당기순이익3 = []
                    for i in arr_new업종_1 :
                        # 해당 업종의 자산총계 구하기
                        try:
                            자산총계 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_자산총계_4'].values[0]
                            if np.isnan(자산총계) :
                                자산총계 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_자산총계(2)_4'].values[0]
                        except:
                            자산총계 = 'non data'
                        
                        arr_업종자산총계.append(자산총계)
                        
                        if 자산총계 != 'non data' :
                            if np.isnan(자산총계) == False :
                                arr_new업종_4.append(i)
                        
                    for i in arr_new업종_4 :
                        # 해당 업종의 당기순이익 구하기
                        try:
                            당기순이익3 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_당기순이익_4'].values[0]
                            if np.isnan(당기순이익3) :
                                당기순이익3 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_당기순이익(손익)_4'].values[0]  
                                if np.isnan(당기순이익3) :
                                    당기순이익3 = data5_4.loc[data5_4['종목명'] == i][str(j)+'_당기순이익(2)_4'].values[0]
                        except:
                            pass
                            
                        arr_업종당기순이익3.append(당기순이익3)    
                            
                    #업종 자본총계 평균 구하기
                    arr_업종자산총계 = [x for x in arr_업종자산총계 if x != 'nan data'] # nan data 값 제거
                    arr_업종자산총계 = [x for x in arr_업종자산총계 if np.isnan(x) == False] # nan 값 제거
                    업종자산총계 = np.mean(arr_업종자산총계)
                    
                    # 업종 ROE 구하기
                    업종_ROA = 업종당기순이익 / 업종자산총계 * 100
                    print('업종당기순이익3: ',arr_업종당기순이익3)
                    print('업종자산총계: ',arr_업종자산총계)
                    print('크기비교(두개가같아야함): ', len(arr_업종당기순이익3), len(arr_업종자산총계))
                    print('업종_ROA : ', 업종_ROA)
                    
                    
                    
                    ##---------------------------------------------------------------------##
                    # 업종의 지표 셀에 넣기
                    sheet.cell(289, 3 + col_num, 업종_PER)
                    sheet.cell(300, 3 + col_num, 업종_PBR)
                    sheet.cell(305, 4 + col_num, 업종_ROE)
                    sheet.cell(311, 4 + col_num, 업종_ROA)
                    
                    
                    
                ##---------------------------------------------------------------------##
                ## 당해 최저가, 최고가 구하기
                당해 = j
                종목명 = name
                path_고저 = "F:/JusikData/oneday_csv/onedaydata/"+name+'/'+name+'.csv'
                data_고저 = pd.read_csv(path_고저, encoding='cp949')
                    
                max_arr = []
                min_arr = []
                for i in data_고저.index:
                    if str(data_고저.iloc[i]['일자'])[0:4] == str(j) :
                        max_arr.append(data_고저.iloc[i]['고가'])
                        min_arr.append(data_고저.iloc[i]['저가'])
                    
                # 고가, 저가에서 0은 제외
                max_arr = [x for x in max_arr if x != 0]
                min_arr = [x for x in min_arr if x != 0]
                    
                # 최대값, 최소값 구하기\
                try:
                    max_price = max(max_arr)
                    min_price = min(min_arr)
                except:
                    max_price = 'non data'
                    min_price = 'non data'
                    
                print('GBN4 - Max : ',max_price)
                print('GBN4 - Min : ',min_price)
                    
                    
                    
                ##---------------------------------------------------------------------##
                # 좌표로 값 출력 = sheet.cell(행, 열, 값)
                sheet.cell(282, 3 + col_num, min_price)
                sheet.cell(283, 3 + col_num, max_price)
                sheet.cell(293, 3 + col_num, min_price)
                sheet.cell(294, 3 + col_num, max_price)
                    
                # 1씩 증가하며 cell에 값넣기 위해서
                col_num += 1
                        
                        
                        
        ##---------------------------------------------------------------------##
        # 저장
        wb.save('F:/JusikData/analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
#conn = GBN4_cls()
#conn.Exe_GBN4('삼성전자',20200504)