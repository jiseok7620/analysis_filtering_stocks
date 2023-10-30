import openpyxl
import pandas as pd
from datetime import datetime
import numpy as np
import mpl_finance
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

class compo_chart_cls:
    def make_roc(self, name, data1, dd):
        # 20일 roc 차트 만들기
        arr_roc_20 = []
        
        # 60봉(3개월)차트로 만들거임
        count_num = 60
        
        # 현재 일자의 인덱스 구하기
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        
        while True:
            if count_num < 0:
                break
            
            당일종가 = data1.iloc[idx-count_num]['종가']
            종가_20 = data1.iloc[idx-count_num-20]['종가']
            
            roc_20 = ((당일종가 - 종가_20) / 종가_20) *100
            
            arr_roc_20.append(roc_20)
            
            count_num -= 1
        
        #print(len(arr_roc_20))
        return arr_roc_20


    
    ##---------------------------------------------------------------------##
    def make_rsi(self, data1, period, dd):
        # 당일 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        
        # RSI 구하기
        U = np.where(np.diff(data1['종가']) > 0, np.diff(data1['종가']), 0)
        D = np.where(np.diff(data1['종가']) < 0, np.diff(data1['종가']) *(-1), 0)
        
        AU = pd.DataFrame(U).rolling(window=period).mean()
        AD = pd.DataFrame(D).rolling(window=period).mean()
        RSI = AU.div(AD+AU) *100
        
        # 변수
        rsi_60 = RSI[idx-61:idx] # 당일인덱스가 RSI의 길이보다 1이 더크므로 +1을 안해줌
        
        # 인덱스 초기화
        rsi_60 = rsi_60.reset_index(drop=True)
        
        return rsi_60
        
        
        
    ##---------------------------------------------------------------------##
    def make_jisu(self, name, data1, dd, how_return):
        # 날짜 형식 바꾸기
        date = str(dd)[0:4] + '-' + str(dd)[4:6] + '-' + str(dd)[6:8]
        
        # 지수를 데이터 프레임으로 가져오기
        kos = data1.iloc[-1]['시장구분']
        #print(kos)
        
        if kos == 'KOSPI' :
            kos_kind = '코스피지수' # or 코스닥지수
        elif kos == 'KOSDAQ':    
            kos_kind = '코스닥지수'
        #print(kos_kind)
        
        # 지수데이터, 종목데이터 가져오기
        jisudata = pd.read_csv("oneday_csv/jisu/" + kos_kind + '.csv', encoding='cp949')
        jongmokdata = pd.read_csv("oneday_csv/onedaydata/"+name+'/'+name+'.csv', encoding='cp949')
        
        # 해당일자의 인덱스 만들기
        
        #print(jisudata)
        #print(date)
        idx = jisudata[jisudata['일자'] == date].index[0]
        idx2 = jongmokdata[jongmokdata['일자'] == dd].index[0]
        
        # 지수와 종목을 넣을 배열만들기
        arr_지수기준 = []
        arr_종목기준 = []
        arr_이격도 = []
        
        # 반복문 카운트
        count_num = 60
        
        while True:
            if count_num < 0:
                break
            
            # 지수 차트 만들기
            지수기준일 = jisudata.iloc[idx-60]['종가']
            지수종가 = jisudata.iloc[idx-count_num]['종가']
            지수선 = 지수종가 / 지수기준일
            지수선_pow = pow(지수선, 1)
            arr_지수기준.append(지수선_pow)

            # 종목 차트 만들기
            종목기준일 = jongmokdata.iloc[idx2-60]['종가']
            종목종가 = jongmokdata.iloc[idx2-count_num]['종가']
            종목선 = 종목종가 / 종목기준일
            종목선_pow = pow(종목선, 1)
            arr_종목기준.append(종목선_pow)
            
            # 이격도 만들기
            지수이격 = (종목선_pow / 지수선_pow) * 100
            arr_이격도.append(지수이격)
            
            count_num -= 1
            
            # 리턴할 나눌값
            nanugi_1 = 종목기준일 / 지수기준일
        
        if how_return == 1:
            return arr_지수기준
        elif how_return == 2:
            return arr_종목기준
        elif how_return == 3:
            return arr_이격도
        elif how_return == 4:
            return nanugi_1
    
    
    
    ##---------------------------------------------------------------------##
    def make_obv(self, name, data1, dd):
        # 현재날짜의 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        
        OBV = []
        OBV.append(0)
        
        for i in range(idx-60,idx+1):
            if data1.iloc[i]['종가'] > data1.iloc[i-1]['종가'] :
                OBV.append(OBV[-1] + data1.iloc[i]['거래량'])
            elif data1.iloc[i]['종가'] < data1.iloc[i-1]['종가'] :
                OBV.append(OBV[-1] - data1.iloc[i]['거래량'])
            else:
                OBV.append(OBV[-1])
        
        # OBV에서 처음에 넣어줬던 0을 삭제하기
        del OBV[OBV.index(0)]
        
        # obv차트 리턴하기
        return OBV



    ##---------------------------------------------------------------------##
    def make_vr(self, name, data1, dd, period):
        # 현재날짜의 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        
        # VR 배열
        VR = []
        
        for i in range(idx-60,idx+1):
            if i >= period :
                # 상승분, 하락분 배열
                VR_up = []
                VR_down = []
                VR_up.append(0)
                VR_down.append(0)
                
                for j in range(i-period+1,i+1):
                    if data1.iloc[j]['종가'] > data1.iloc[j-1]['종가'] :
                        VR_up.append(VR_up[-1] + data1.iloc[j]['거래량'])
                        VR_down.append(VR_down[-1])
                    elif data1.iloc[j]['종가'] < data1.iloc[j-1]['종가'] :
                        VR_up.append(VR_up[-1])
                        VR_down.append(VR_down[-1] + data1.iloc[j]['거래량'])
                    else:
                        VR_up.append(VR_up[-1] + data1.iloc[j]['거래량']/2)
                        VR_down.append(VR_down[-1] + data1.iloc[j]['거래량']/2)
                
                VR.append((VR_up[-1] / VR_down[-1]) *100)
                
            else :
                VR.append(0)
        
        # 리턴
        return VR

            
    ##---------------------------------------------------------------------##
    def Exe_compo_chart(self,name,dd):
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']
        sheet2 = wb['종합차트']



        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        #print(len(data1))
        
        
        
        ##---------------------------------------------------------------------##
        # 기본차트만들기
        # 클래스 객체 생성
        conn = compo_chart_cls()
        '''
        # 날짜 형식 바꾸기
        date = str(dd)[0:4] + '-' + str(dd)[4:6] + '-' + str(dd)[6:8]
        
        # 지수를 데이터 프레임으로 가져오기
        kos = data1.iloc[-1]['시장구분']
        #print(kos)
        
        if kos == 'KOSPI' :
            kos_kind = '코스피지수' # or 코스닥지수
        elif kos == 'KOSDAQ':    
            kos_kind = '코스닥지수'
        #print(kos_kind)
        
        # 지수데이터, 종목데이터 가져오기
        jisudata = pd.read_csv("F:/JusikData/oneday_csv/jisu/" + kos_kind + '.csv', encoding='cp949')
        
        # 현재날짜의 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        idx2 = jisudata[jisudata['일자'] == date].index[0]
        
        # 지수기준일, 종목기준일 나눌값
        nanugi_1 = conn.make_jisu(name, data1, dd, 4)
        
        # 시고저종 거래량 가져오기
        y_data_close = data1['종가'].values.tolist() / nanugi_1
        y_data_start = data1['시가'].values.tolist() / nanugi_1
        y_data_high = data1['고가'].values.tolist() / nanugi_1
        y_data_low = data1['저가'].values.tolist() / nanugi_1
        y_data_trade = data1['거래량'].values.tolist() / nanugi_1
        
        # 이동평균만들기
        ma5 = data1['종가'].rolling(window=5).mean()
        ma10 = data1['종가'].rolling(window=10).mean()
        ma20 = data1['종가'].rolling(window=20).mean()
        
        # 60봉 동안의 데이터
        y_data_close = y_data_close[idx-60:idx+1] # 종가
        y_data_start = y_data_start[idx-60:idx+1] # 시가
        y_data_high = y_data_high[idx-60:idx+1] # 고가
        y_data_low = y_data_low[idx-60:idx+1] # 저가
        y_data_trade = y_data_trade[idx-60:idx+1] # 거래량
        
        # 60봉 동안의 이동평균데이터
        avgline_5 = ma5[idx-60:idx+1]
        avgline_10 = ma10[idx-60:idx+1]
        avgline_20 = ma20[idx-60:idx+1]
        
        # 지수의 봉 차트구하기위해 시고저종 구하기
        y_data_close_2 = jisudata['종가'].values.tolist()
        y_data_start_2 = jisudata['시가'].values.tolist()
        y_data_high_2 = jisudata['고가'].values.tolist()
        y_data_low_2 = jisudata['저가'].values.tolist()
        
        y_data_close_2 = y_data_close_2[idx2-60:idx2+1] # 종가
        y_data_start_2 = y_data_start_2[idx2-60:idx2+1] # 시가
        y_data_high_2 = y_data_high_2[idx2-60:idx2+1] # 고가
        y_data_low_2 = y_data_low_2[idx2-60:idx2+1] # 저가
        '''
        
        
        
        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        # ROC 그래프 데이터프레임 가져오기
        ROC_20 = conn.make_roc(name, data1, dd)
        
        # RSI 그래프 데이터프레임 가져오기
        RSI_20 = conn.make_rsi(data1, 60, dd)
        
        # 지수-종목 3개월 데이터 프레임 가져오기
        jisu_60 = conn.make_jisu(name, data1, dd, 1)
        jongmok_60 = conn.make_jisu(name, data1, dd, 2)
        
        # 지수-종목 이격도
        ji_jong_이격 = conn.make_jisu(name, data1, dd, 3)
        
        # obv 차트
        obv_60 = conn.make_obv(name, data1, dd)
        
        # vr 차트
        vr_60 = conn.make_vr(name, data1, dd, 20)
        
        # 크기보기
        #print(len(ROC_20))
        #print(len(RSI_20))        
        #print(len(jisu_60))
        #print(len(jongmok_60))
        #print(len(ji_jong_이격))
        #print(len(obv_60))
        #print(len(vr_60))
        
        '''
        ##---------------------------------------------------------------------##
        # 그래프 그리기
        # 캔버스 설정(크기, 배경 설정 등)
        fig = plt.figure(figsize=(40,25)) ## 캔버스 생성
        fig.set_facecolor('white') ## 캔버스 색상 설정
        
        # 그림 뼈대(프레임) 생성 
        #차트1 : 지수-종목그래프
        ax1 = fig.add_subplot(421)
        ax1.plot(jisu_60, color='black', marker='o', markersize=3, label='jisu')
        ax1.plot(jongmok_60, color='red', marker='o', markersize=3, label='jong')
        plt.xticks(visible=False) # 축값없애기
        plt.legend() # 범례만들기
        
        #차트8 : 지수-종목 이격도 바차트
        ax8 = fig.add_subplot(422)
        plt.bar(range(61), ji_jong_이격)
        
        #차트2 : 3개월 일봉차트
        ax2 = fig.add_subplot(423)
        mpl_finance.candlestick2_ohlc(ax2, y_data_start, y_data_high, y_data_low, y_data_close, width=0.5, colorup='r', colordown='b')
        mpl_finance.candlestick2_ohlc(ax2, y_data_start_2, y_data_high_2, y_data_low_2, y_data_close_2, width=0.5, colorup='green', colordown='black')
        plt.xticks(visible=False) # 축값없애기
        plt.grid(True, axis='y') # 그리드(격자)
        # 이동평균 추가하기
        #plt.plot(range(61), avgline_5, 'g', label='5') # 5일 이동평균선 그리기
        #plt.plot(range(61), avgline_10, 'b', label='10') # 차트 그리기
        #plt.plot(range(61), avgline_20, 'r', label='20') # 차트 그리기
        
        # 차트7 : 거래량차트
        # 거래량 바 차트
        ax7 = fig.add_subplot(424)
        plt.bar(range(61), y_data_trade)
        
        # 차트3 : ROC차트
        ax3 = fig.add_subplot(425)
        ax3.plot(ROC_20,color='black', marker='o', markersize=3, label='ROC')
        plt.xticks(visible=False) # 축값없애기
        plt.legend() # 범례만들기
        
        # 차트4 : RSI차트
        ax4 = fig.add_subplot(426)
        ax4.plot(RSI_20, color='black', marker='o', markersize=3, label='RSI')
        plt.xticks(visible=False) # 축값없애기
        plt.legend() # 범례만들기
        
        
        # 차트5 : OBV 차트
        ax5 = fig.add_subplot(427)
        ax5.plot(obv_60[0:60], color='black', marker='o', markersize=3, label='obv')
        plt.xticks(visible=False) # 축값없애기
        plt.legend() # 범례만들기
        
        # 차트6 : VR차트
        ax6 = fig.add_subplot(428)
        ax6.plot(vr_60[0:60], color='black', marker='o', markersize=3, label='vr')
        plt.legend() # 범례만들기
        
        # 수평선 긋기
        ax1.axhline(1,color='green') # 지수-종목차트
        ax3.axhline(0,color='green') # ROC 차트
        ax4.axhline(30,color='green') # RSI차트
        ax4.axhline(50,color='green')
        ax4.axhline(70,color='green')
        ax6.axhline(100,color='green') # VR차트
        ax6.axhline(150,color='green')
        
        # 그래프 보기
        #plt.show()
        
        # 저장
        plt.savefig('F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_종합차트.png')
        
        # 메모리 제거
        plt.close()
        
        
        
        ##---------------------------------------------------------------------##
        # 그래프 붙여넣기
        #이미지 경로지정, 이미지 붙이기
        img_path = 'F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_종합차트.png'
        img = Image(img_path)
        sheet2.add_image(img,'A1')
        
        
        '''
        ##---------------------------------------------------------------------##
        ## 지표값 넣기
        count_col = 0 # 열
        count_num = 0 # 반복수
        
        while True:
            # 정지
            if count_num == 61 : 
                break
            
            # 0째가 제일 최근 60번째가 제일 나중에
            ROC = ROC_20[count_num]
            RSI = RSI_20.iloc[count_num][0]
            OBV = obv_60[count_num]
            VR = vr_60[count_num]

            # 반복
            count_num += 1
            
            ##---------------------------------------------------------------------##
            # 셀에 값입력
            sheet.cell(135, 3+count_col, ROC)
            sheet.cell(136, 3+count_col, RSI)
            sheet.cell(137, 3+count_col, OBV)
            sheet.cell(141, 3+count_col, VR)
            
            # 열 이동
            count_col += 1
        
        ##---------------------------------------------------------------------##        
        # 저장
        wb.save('analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
#conn = compo_chart_cls()
#conn.Exe_compo_chart('셀트리온',20211209)