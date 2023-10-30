import openpyxl
import pandas as pd
import numpy as np
from Invest.analysis_excel.except_hol import except_hol_cls

class jaemu_info_cls:
    # name = 종목명, dd = 일자, q = 분기
    def Exe_jaemu_info(self,name,dd, q):
        ##---------------------------------------------------------------------##
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']
        
        
        
        ##---------------------------------------------------------------------##
        ## 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        #print(data1.columns)
        
        ## data2 = 재무 정보 데이터 ##
        path2_1 = "report_csv/report_make/재무정보_1분기.csv"
        path2_2 = "report_csv/report_make/재무정보_2분기.csv"
        path2_3 = "report_csv/report_make/재무정보_3분기.csv"
        path2_4 = "report_csv/report_make/재무정보_4분기.csv"
        data2_1 = pd.read_csv(path2_1, encoding='cp949')
        data2_2 = pd.read_csv(path2_2, encoding='cp949')
        data2_3 = pd.read_csv(path2_3, encoding='cp949')
        data2_4 = pd.read_csv(path2_4, encoding='cp949')
        #print(data2_1.columns)
        
        
        
        ##---------------------------------------------------------------------##
        ## 변수에 데이터 저장
        # 당해 구하기(과거 데이터 분석 시)
        dd_str = str(dd)[0:4]
        dd_int = int(dd_str)
        
        # dd의 년, 월, 일 나누기
        dd_y = int(str(dd)[0:4])
        dd_m = int(str(dd)[4:6])
        dd_d = int(str(dd)[6:8])
        
        # 년도, 분기 배열 만들기
        arr_yago = [dd_y,dd_y-1,dd_y-2,dd_y-3]
        arr_q = [1,2,3,4]
        row_num_1 = 0
        col_num_1 = 0
        
        # 년도, 분기 반복문 돌리기        
        for i in arr_yago:
            for k in arr_q:
                if k == 1:
                    data = data2_1
                elif k == 2:
                    data = data2_2
                elif k == 3:
                    data = data2_3
                elif k == 4:
                    data = data2_4
                
                # 데이터를 변수에 넣기
                # 자산총액
                try:
                    자산총액 = data.loc[data['종목명'] == name][str(i)+'_자산총계_'+str(k)].values[0]
                    if np.isnan(자산총액) :
                        자산총액 = data.loc[data['종목명'] == name][str(i)+'_자산총계(2)_'+str(k)].values[0]
                except:
                    자산총액 = 'non data'
                    
                # 유동자산
                try:
                    유동자산 = data.loc[data['종목명'] == name][str(i)+'_유동자산_'+str(k)].values[0]
                    if np.isnan(유동자산) :
                        유동자산 = data.loc[data['종목명'] == name][str(i)+'_유동자산(2)_'+str(k)].values[0]
                except:
                    유동자산 = 'non data'
                    
                # 매출채권
                try:
                    매출채권 = data.loc[data['종목명'] == name][str(i)+'_매출채권(유동)_'+str(k)].values[0]
                    if np.isnan(매출채권) :
                        매출채권 = data.loc[data['종목명'] == name][str(i)+'_매출채권(유동)(2)_'+str(k)].values[0]
                except:
                    매출채권 = 'non data'
                
                # 재고자산
                try:
                    재고자산 = data.loc[data['종목명'] == name][str(i)+'_재고자산_'+str(k)].values[0]
                    if np.isnan(재고자산) :
                        재고자산 = data.loc[data['종목명'] == name][str(i)+'_재고자산(2)_'+str(k)].values[0]
                except:
                    재고자산 = 'non data'
                    
                # 부채총액
                try:
                    부채총액 = data.loc[data['종목명'] == name][str(i)+'_부채총계_'+str(k)].values[0]
                    if np.isnan(부채총액) :
                        부채총액 = data.loc[data['종목명'] == name][str(i)+'_부채총계(2)_'+str(k)].values[0]
                except:
                    부채총액 = 'non data'
                
                # 유동부채
                try:
                    유동부채 = data.loc[data['종목명'] == name][str(i)+'_유동부채_'+str(k)].values[0]
                    if np.isnan(유동부채) :
                        유동부채 = data.loc[data['종목명'] == name][str(i)+'_유동부채(2)_'+str(k)].values[0]
                except:
                    유동부채 = 'non data'
                    
                # 전환사채
                try:
                    전환사채 = data.loc[data['종목명'] == name][str(i)+'_전환사채(유동)_'+str(k)].values[0]
                    if np.isnan(전환사채) :
                        전환사채 = data.loc[data['종목명'] == name][str(i)+'_전환사채(유동)(2)_'+str(k)].values[0]
                except:
                    전환사채 = 'non data'
                    
                # 신주인수권부사채
                try:
                    신주인수권부사채 = data.loc[data['종목명'] == name][str(i)+'_신주인수권부사채(유동)_'+str(k)].values[0]
                    if np.isnan(신주인수권부사채) :
                        신주인수권부사채 = data.loc[data['종목명'] == name][str(i)+'_신주인수권부사채(유동)(2)_'+str(k)].values[0]
                except:
                    신주인수권부사채 = 'non data'
                
                # 자본총액
                try:
                    자본총액 = data.loc[data['종목명'] == name][str(i)+'_자본총계_'+str(k)].values[0]
                    if np.isnan(자본총액) :
                        자본총액 = data.loc[data['종목명'] == name][str(i)+'_자본총계(2)_'+str(k)].values[0]
                except:
                    자본총액 = 'non data'
                
                # 자본금
                try:
                    자본금 = data.loc[data['종목명'] == name][str(i)+'_자본금_'+str(k)].values[0]
                    if np.isnan(자본금) :
                        자본금 = data.loc[data['종목명'] == name][str(i)+'_자본금(2)_'+str(k)].values[0]
                except:
                    자본금 = 'non data'
                    
                # 자본잉여금
                try:
                    자본잉여금 = data.loc[data['종목명'] == name][str(i)+'_자본잉여금_'+str(k)].values[0]
                    if np.isnan(자본잉여금) :
                        자본잉여금 = data.loc[data['종목명'] == name][str(i)+'_자본잉여금(2)_'+str(k)].values[0]
                except:
                    자본잉여금 = 'non data'
                    
                # 이익잉여금
                try:
                    이익잉여금 = data.loc[data['종목명'] == name][str(i)+'_이익잉여금_'+str(k)].values[0]
                    if np.isnan(이익잉여금) :
                        이익잉여금 = data.loc[data['종목명'] == name][str(i)+'_이익잉여금(2)_'+str(k)].values[0]
                except:
                    이익잉여금 = 'non data'
                    
                # 매출액
                try:
                    매출액 = data.loc[data['종목명'] == name][str(i)+'_매출액_'+str(k)].values[0]
                    if np.isnan(매출액) :
                        매출액 = data.loc[data['종목명'] == name][str(i)+'_매출액(손익)_'+str(k)].values[0]
                        if np.isnan(매출액) :
                            매출액 = data.loc[data['종목명'] == name][str(i)+'_매출액(2)_'+str(k)].values[0]                        
                except:
                    매출액 = 'non data'
                    
                # 매출원가
                try:
                    매출원가 = data.loc[data['종목명'] == name][str(i)+'_매출원가_'+str(k)].values[0]
                    if np.isnan(매출원가) :
                        매출원가 = data.loc[data['종목명'] == name][str(i)+'_매출원가(손익)_'+str(k)].values[0]
                        if np.isnan(매출원가) :
                            매출원가 = data.loc[data['종목명'] == name][str(i)+'_매출원가(2)_'+str(k)].values[0]                        
                except:
                    매출원가 = 'non data'
                    
                # 영업이익
                try:
                    영업이익 = data.loc[data['종목명'] == name][str(i)+'_영업이익_'+str(k)].values[0]
                    if np.isnan(영업이익) :
                        영업이익 = data.loc[data['종목명'] == name][str(i)+'_영업이익(손익)_'+str(k)].values[0]
                        if np.isnan(영업이익) :
                            영업이익 = data.loc[data['종목명'] == name][str(i)+'_영업이익(2)_'+str(k)].values[0]                      
                except:
                    영업이익 = 'non data'
                    
                # 당기순이익
                try:
                    당기순이익 = data.loc[data['종목명'] == name][str(i)+'_당기순이익_'+str(k)].values[0]
                    if np.isnan(당기순이익) :
                        당기순이익 = data.loc[data['종목명'] == name][str(i)+'_당기순이익(손익)_'+str(k)].values[0]  
                        if np.isnan(당기순이익) :
                            당기순이익 = data.loc[data['종목명'] == name][str(i)+'_당기순이익(2)_'+str(k)].values[0]                      
                except:
                    당기순이익 = 'non data'
                    
                # 비지배지분이익
                try:
                    비지배지분이익 = data.loc[data['종목명'] == name][str(i)+'_비지배지분이익_'+str(k)].values[0]
                    if np.isnan(비지배지분이익) :
                        비지배지분이익 = data.loc[data['종목명'] == name][str(i)+'_비지배지분이익(손익)_'+str(k)].values[0]          
                        if np.isnan(비지배지분이익) :
                            비지배지분이익 = data.loc[data['종목명'] == name][str(i)+'_비지배지분이익(2)_'+str(k)].values[0]                 
                except:
                    비지배지분이익 = 'non data'
                    
                # 영업활동현금흐름
                try:
                    영업활동현금흐름 = data.loc[data['종목명'] == name][str(i)+'_영업활동현금흐름_'+str(k)].values[0]
                    if np.isnan(영업활동현금흐름) :
                        영업활동현금흐름 = data.loc[data['종목명'] == name][str(i)+'_영업활동현금흐름(2)_'+str(k)].values[0] 
                except:
                    영업활동현금흐름 = 'non data'
                    
                # 투자활동현금흐름
                try:
                    투자활동현금흐름 = data.loc[data['종목명'] == name][str(i)+'_투자활동현금흐름_'+str(k)].values[0]
                    if np.isnan(투자활동현금흐름) :
                        투자활동현금흐름 = data.loc[data['종목명'] == name][str(i)+'_투자활동현금흐름(2)_'+str(k)].values[0]
                except:
                    투자활동현금흐름 = 'non data'
                
                # 재무활동현금흐름
                try:
                    재무활동현금흐름 = data.loc[data['종목명'] == name][str(i)+'_재무활동현금흐름_'+str(k)].values[0]
                    if np.isnan(재무활동현금흐름) :
                        재무활동현금흐름 = data.loc[data['종목명'] == name][str(i)+'_재무활동현금흐름(2)_'+str(k)].values[0]
                except:
                    재무활동현금흐름 = 'non data'

                ## 발행주식수(1분기 : 03.29 / 2분기 : 06.29 / 3분기 : 09.29 / 4분기 : 12.29) 직접 수기 작성필요 -> 수정하기
                #발행주식수 
                # if문을 이용해서 1~4분기의 날짜가 토요일이면 -1, 일요일이면 +1하기
                dd_1 = int(str(i)[0:4]+'03'+'29')
                dd_2 = int(str(i)[0:4]+'06'+'29')
                dd_3 = int(str(i)[0:4]+'09'+'29')
                dd_4 = int(str(i)[0:4]+'12'+'29')
                
                # 분기에 따라서 발행주식수 계산
                if k == 1:
                    dd_jusiksu = except_hol_cls.exe_except(self, dd_1)
                elif k == 2:
                    dd_jusiksu = except_hol_cls.exe_except(self, dd_2)
                elif k == 3:
                    dd_jusiksu = except_hol_cls.exe_except(self, dd_3)
                elif k == 4:
                    dd_jusiksu = except_hol_cls.exe_except(self, dd_4)
                
                # 해당 분기 전까지 데이터만
                if i == dd_y and k > q:
                    break
                
                #print(dd_jusiksu)
                
                try:
                    상장주식수 = data1.loc[data1['일자'] == int(dd_jusiksu)]['상장주식수'].values[0]
                except:
                    상장주식수 = 'non data'
                
                
                
                ##---------------------------------------------------------------------##    
                # 변수의 데이터를 cell에 넣는 구문넣기
                # 열 간격이 2, 행 간격이 1
                sheet.cell(11 + row_num_1, 4 + col_num_1, 자산총액)
                sheet.cell(15 + row_num_1, 4 + col_num_1, 유동자산)
                sheet.cell(19 + row_num_1, 4 + col_num_1, 매출채권)
                sheet.cell(23 + row_num_1, 4 + col_num_1, 재고자산)
                sheet.cell(29 + row_num_1, 4 + col_num_1, 부채총액)
                sheet.cell(33 + row_num_1, 4 + col_num_1, 유동부채)
                sheet.cell(37 + row_num_1, 4 + col_num_1, 전환사채)
                sheet.cell(41 + row_num_1, 4 + col_num_1, 신주인수권부사채)
                sheet.cell(45 + row_num_1, 4 + col_num_1, 상장주식수)
                sheet.cell(51 + row_num_1, 4 + col_num_1, 자본총액)
                sheet.cell(55 + row_num_1, 4 + col_num_1, 자본금)
                sheet.cell(59 + row_num_1, 4 + col_num_1, 자본잉여금)
                sheet.cell(63 + row_num_1, 4 + col_num_1, 이익잉여금)
                sheet.cell(73 + row_num_1, 4 + col_num_1, 매출액)
                sheet.cell(77 + row_num_1, 4 + col_num_1, 매출원가)
                sheet.cell(83 + row_num_1, 4 + col_num_1, 영업이익)
                sheet.cell(89 + row_num_1, 4 + col_num_1, 당기순이익)
                sheet.cell(93 + row_num_1, 4 + col_num_1, 비지배지분이익)
                sheet.cell(99 + row_num_1, 4 + col_num_1, 영업활동현금흐름)
                sheet.cell(103 + row_num_1, 4 + col_num_1, 투자활동현금흐름)
                sheet.cell(107 + row_num_1, 4 + col_num_1, 재무활동현금흐름)
                
                # 행은 1씩 증가
                row_num_1 += 1
            
            # 다시 행번호를 초기화
            row_num_1 = 0
            # 열은 1씩 증가
            col_num_1 += 1
            
            
            
        ##---------------------------------------------------------------------##
        # 저장
        wb.save('analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
#conn = GBN2_cls()
#conn.Exe_GBN2('삼성전자',20200504)