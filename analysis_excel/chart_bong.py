import openpyxl
import pandas as pd
from datetime import datetime
import mplfinance as mpf
from openpyxl.drawing.image import Image

class bong_chart_cls:
    def Exe_bong_chart(self,name,dd):
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']


        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        #print(data1.columns)
        
        
        
        ##---------------------------------------------------------------------##
        ## 변수에 데이터 저장
        # 당일 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        
        # 일봉의 시고저종, 거래량
        일봉_고가 = data1.iloc[idx]['고가']
        일봉_종가 = data1.iloc[idx]['종가']
        일봉_저가 = data1.iloc[idx]['저가']
        일봉_시가 = data1.iloc[idx]['시가']

        # 주봉(5개)의 시고저종, 거래량
        주봉_고가 = data1[idx-4:idx+1]['고가'].max()
        주봉_시가 = data1.iloc[idx-4]['시가']
        # 저가(5개)를 배열로 만들기
        주봉_저가_arr = data1[idx-4:idx+1]['저가']
        # 저가(5개)에서 0은 제외
        주봉_저가_arr = [x for x in 주봉_저가_arr if x != 0]
        # 주가저가(5개)구하기
        주봉_저가 = min(주봉_저가_arr)

        # 월봉(20개)의 시고저종, 거래량
        월봉_고가 = data1[idx-19:idx+1]['고가'].max()
        월봉_시가 = data1.iloc[idx-19]['시가']
        # 저가(20개)를 배열로 만들기
        월봉_저가_arr = data1[idx-19:idx+1]['저가']
        # 저가(20개)에서 0은 제외
        월봉_저가_arr = [x for x in 월봉_저가_arr if x != 0]
        # 주가저가(20개)구하기
        월봉_저가 = min(월봉_저가_arr)

        # 3개월봉(60개)의 시고저종, 거래량
        삼개월봉_고가 = data1[idx-59:idx+1]['고가'].max()
        삼개월봉_시가 = data1.iloc[idx-59]['시가']
        # 저가(60개)를 배열로 만들기
        삼개월봉_저가_arr = data1[idx-59:idx+1]['저가']
        print(삼개월봉_저가_arr)
        # 저가(60개)에서 0은 제외
        삼개월봉_저가_arr = [x for x in 삼개월봉_저가_arr if x != 0]
        print(삼개월봉_저가_arr)
        # 주가저가(60개)구하기
        삼개월봉_저가 = min(삼개월봉_저가_arr)

        # 6개월봉(120개)의 시고저종, 거래량
        육개월봉_고가 = data1[idx-119:idx+1]['고가'].max()
        육개월봉_시가 = data1.iloc[idx-119]['시가']
        # 저가(120개)를 배열로 만들기
        육개월봉_저가_arr = data1[idx-119:idx+1]['저가']
        # 저가(120개)에서 0은 제외
        육개월봉_저가_arr = [x for x in 육개월봉_저가_arr if x != 0]
        # 주가저가(120개)구하기
        육개월봉_저가 = min(육개월봉_저가_arr)

        # 1년봉(240개)의 시고저종, 거래량
        일년봉_고가 = data1[idx-239:idx+1]['고가'].max()
        일년봉_시가 = data1.iloc[idx-239]['시가']
        # 저가(240개)를 배열로 만들기
        일년봉_저가_arr = data1[idx-239:idx+1]['저가']
        # 저가(240개)에서 0은 제외
        일년봉_저가_arr = [x for x in 일년봉_저가_arr if x != 0]
        # 주가저가(240개)구하기
        일년봉_저가 = min(일년봉_저가_arr)
        '''
        # 일자를 datetime 형식으로 바꾼뒤 배열에 저장하기
        sl_date = str(dd)[0:4] + '-' + str(dd)[4:6] + '-' + str(dd)[6:8]
        real_date = datetime.strptime(sl_date, "%Y-%m-%d").date()
        
        # 데이터프레임 만들기
        bong_pd = pd.DataFrame({'Date': [real_date,real_date,real_date,real_date,real_date,real_date],
                            'Open':[일봉_시가, 주봉_시가, 월봉_시가, 삼개월봉_시가, 육개월봉_시가, 일년봉_시가],
                            'High':[일봉_고가, 주봉_고가, 월봉_고가, 삼개월봉_고가, 육개월봉_고가, 일년봉_고가],
                            'Low':[일봉_저가, 주봉_저가, 월봉_저가, 삼개월봉_저가, 육개월봉_저가, 일년봉_저가],
                            'Close':[일봉_종가, 일봉_종가, 일봉_종가, 일봉_종가, 일봉_종가, 일봉_종가]})
        
        
        
        ##---------------------------------------------------------------------##
        ## 봉 차트 만들기
        # 일자를 datetime형식으로 바꾸고 인덱스로 설정하기
        bong_pd['Date'] = pd.to_datetime(bong_pd['Date'].values.tolist())
        bong_pd = bong_pd.set_index("Date")
        #print(bong_pd)
        
        # 차트 설정
        colorset = mpf.make_marketcolors(up='tab:red', down='tab:blue', volume='tab:blue')
        s = mpf.make_mpf_style(base_mpf_style='default', marketcolors=colorset)
        
        # 그래프의 시작과 끝 정하기
        arr_start = [0,1,2,3,4,5]
        
        # 그래프 그리기
        mpf.plot(bong_pd.iloc[0:6],            # 데이터 지정
                figscale=0.8,          # 그림 크기 지정
                figratio=(15,5),       # 그림 높이, 너비 비율 지정
                savefig=dict(fname='F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_봉차트.png',dpi=100,pad_inches=0.25),    # 저장
                type='candle',         # 차트 종류 지정
                style=s)  
        
        #이미지 경로지정
        img_path = 'F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_봉차트.png'
        img = Image(img_path)
        sheet.add_image(img,'AG4')
        '''
        
        ##---------------------------------------------------------------------##
        # 좌표로 값 출력 = sheet.cell(행, 열, 값)
        if 일봉_종가 >= 일봉_시가 : # 양봉이다
            일봉_위꼬리 = 일봉_고가 - 일봉_종가
            일봉_몸통 = 일봉_종가 - 일봉_시가
            일봉_아래꼬리 = 일봉_시가 - 일봉_저가
        else: # 음봉이다
            일봉_위꼬리 = 일봉_고가 - 일봉_시가
            일봉_몸통 = 일봉_종가 - 일봉_시가
            일봉_아래꼬리 = 일봉_종가 - 일봉_저가
            
        if 일봉_종가 >= 주봉_시가 : # 양봉이다
            주봉_위꼬리 = 주봉_고가 - 일봉_종가
            주봉_몸통 = 일봉_종가 - 주봉_시가
            주봉_아래꼬리 = 주봉_시가 - 주봉_저가
        else: # 음봉이다
            주봉_위꼬리 = 주봉_고가 - 주봉_시가
            주봉_몸통 = 일봉_종가 - 주봉_시가
            주봉_아래꼬리 = 일봉_종가 - 주봉_저가
            
        if 일봉_종가 >= 월봉_시가 : # 양봉이다
            월봉_위꼬리 = 월봉_고가 - 일봉_종가
            월봉_몸통 = 일봉_종가 - 월봉_시가
            월봉_아래꼬리 = 월봉_시가 - 월봉_저가
        else: # 음봉이다
            월봉_위꼬리 = 월봉_고가 - 월봉_시가
            월봉_몸통 = 일봉_종가 - 월봉_시가
            월봉_아래꼬리 = 일봉_종가 - 월봉_저가
            
        if 일봉_종가 >= 일년봉_시가 : # 양봉이다
            년봉_위꼬리 = 일년봉_고가 - 일봉_종가
            년봉_몸통 = 일봉_종가 - 일년봉_시가
            년봉_아래꼬리 = 일년봉_시가 - 일년봉_저가
        else: # 음봉이다
            년봉_위꼬리 = 일년봉_고가 - 일년봉_시가
            년봉_몸통 = 일봉_종가 - 일년봉_시가
            년봉_아래꼬리 = 일봉_종가 - 일년봉_저가
                
        sheet.cell(25,27,일봉_위꼬리)
        sheet.cell(26,27,일봉_몸통)
        sheet.cell(27,27,일봉_아래꼬리)
        sheet.cell(25,28,주봉_위꼬리)
        sheet.cell(26,28,주봉_몸통)
        sheet.cell(27,28,주봉_아래꼬리)
        sheet.cell(25,29,월봉_위꼬리)
        sheet.cell(26,29,월봉_몸통)
        sheet.cell(27,29,월봉_아래꼬리)           
        sheet.cell(25,30,년봉_위꼬리)
        sheet.cell(26,30,년봉_몸통)
        sheet.cell(27,30,년봉_아래꼬리)          

        ##---------------------------------------------------------------------##
        # 저장
        wb.save('analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
#conn = GBA1_cls()
#conn.Exe_GBA1('삼성전자',20200504)