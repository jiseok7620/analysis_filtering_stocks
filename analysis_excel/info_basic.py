import openpyxl
import pandas as pd

class basic_info_cls:
    # name = 종목명, dd = 일자, q = 보고서 유무 상 q, 실제 분기 = q_num
    def Exe_basic_info(self,name, dd, q):
        ##---------------------------------------------------------------------##
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('analysis_csv/HJS/format/개별분석양식(2).xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']



        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        #print(data1.columns)
        
        ## data4 = 업종 분류 데이터 ##
        path4 = "oneday_csv/data_sectors/"+str(dd)[0:4]+'.csv'
        data4 = pd.read_csv(path4, encoding='cp949')
        #print(data4.columns)
        
        
        
        ##---------------------------------------------------------------------##
        ## 변수에 데이터 저장
        # 해당일의 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        
        # 변수에 데이터 저장
        매수일자 = dd
        소속부 = data1.iloc[idx]['소속부']
        종목명 = name
        분기 = q 
        #print('분기는 : ', q)
        상장주식수 = data1.iloc[idx]['상장주식수']
        try:
            업종구분 = data4.loc[data4['종목명'] == name]['업종명'].values[0]
        except:
            업종구분 = 'non data'
        상장시장 = data1.iloc[idx]['시장구분']
        시가총액 = data1.iloc[idx]['시가총액']
        종가 = data1.iloc[idx]['종가']
        거래량 = data1.iloc[idx]['거래량']
        거래대금 = data1.iloc[idx]['거래대금']
        시총 = data1.iloc[idx]['시가총액']
        시총_30 = data1.iloc[idx-30]['시가총액']
        
        
        
        ##---------------------------------------------------------------------##
        # 좌표로 값 출력 = sheet.cell(행, 열, 값)
        sheet.cell(3, 3, 매수일자)
        sheet.cell(3, 5, 분기)
        sheet.cell(4, 3, 종목명)
        sheet.cell(4, 7, 소속부)
        sheet.cell(4, 11, 업종구분)
        sheet.cell(4, 14, 상장시장)
        sheet.cell(5, 3, 종가)
        sheet.cell(5, 5, 상장주식수)
        sheet.cell(5, 8, 시가총액)
        sheet.cell(5, 11, 거래량)
        sheet.cell(5, 14, 거래대금)
        sheet.cell(22, 11, 시총)
        sheet.cell(22, 12, 시총_30)
        sheet.cell(4, 17, "https://finance.naver.com/item/fchart.naver?code="+str(data1.iloc[1]['종목코드']))
        
        ##---------------------------------------------------------------------##
        # 저장
        wb.save('analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
#conn = basic_info_cls()
#conn.Exe_basic_info('삼성전자',20200504,1)