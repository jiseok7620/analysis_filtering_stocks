import openpyxl
import pandas as pd

class chart_info_cls:
    def exe_chart_info(self,name,dd):
        ##---------------------------------------------------------------------##
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']



        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        
        # 이평선 값 추가
        ma5 = data1['종가'].rolling(window=5).mean()
        ma10 = data1['종가'].rolling(window=10).mean()
        ma20 = data1['종가'].rolling(window=20).mean()
        ma60 = data1['종가'].rolling(window=60).mean()
        ma120 = data1['종가'].rolling(window=120).mean()
        ma240 = data1['종가'].rolling(window=240).mean()
        
        data1['5이평'] = ma5
        data1['10이평'] = ma10
        data1['20이평'] = ma20        
        data1['60이평'] = ma60
        data1['120이평'] = ma120
        data1['240이평'] = ma240
                
        ## 지수데이터
        kos = data1.iloc[-1]['시장구분']
        
        if kos == 'KOSPI' :
            kos_kind = '코스피지수' # or 코스닥지수
        elif kos == 'KOSDAQ':    
            kos_kind = '코스닥지수'
        
        jisudata = pd.read_csv("oneday_csv/jisu/" + kos_kind + '.csv', encoding='cp949')
        
        ## data3 = 수급 데이터 ##
        try:
            path3 = "oneday_csv/onedaydata_supply/"+name+'.csv'
            data3 = pd.read_csv(path3, encoding='cp949')
        except:
            pass

        
        ##---------------------------------------------------------------------##
        # 날짜 형식 바꾸기
        date = str(dd)[0:4] + '-' + str(dd)[4:6] + '-' + str(dd)[6:8]
        
        # data3의 일자 형식 바꾸기
        dd_y = int(str(dd)[0:4])
        dd_m = int(str(dd)[4:6])
        dd_d = int(str(dd)[6:8])
        if len(str(dd_m)) == 1:
            if len(str(dd_d)) == 1:
                dd_full = str(dd_y)[0:4] + "/0" + str(dd_m) + "/0" + str(dd_d)
            else:
                dd_full = str(dd_y)[0:4] + "/0" + str(dd_m) + "/" + str(dd_d)
        else :
            if len(str(dd_d)) == 1:
                dd_full = str(dd_y)[0:4] + "/" + str(dd_m) + "/0" + str(dd_d)
            else:
                dd_full = str(dd_y)[0:4] + "/" + str(dd_m) + "/" + str(dd_d)
        
        # 오늘날 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        idx_jisu = jisudata[jisudata['일자'] == date].index[0]
        try:
            idx_3 = data3.loc[data3['일자'] == dd_full]['일자'].index[0]
        except:
            pass
            
        count_col = 0 # 열
        count_num = 60 # 반복수
        
        while True:
            # 정지
            if count_num == -1 : 
                break
            
            # 변수 지정
            지수시가 = jisudata.iloc[idx_jisu-count_num]['시가']
            지수고가 = jisudata.iloc[idx_jisu-count_num]['고가']
            지수저가 = jisudata.iloc[idx_jisu-count_num]['저가']
            지수종가 = jisudata.iloc[idx_jisu-count_num]['종가']
            시가 = data1.iloc[idx-count_num]['시가']
            고가 = data1.iloc[idx-count_num]['고가']
            저가 = data1.iloc[idx-count_num]['저가']
            종가 = data1.iloc[idx-count_num]['종가']
            거래량 = data1.iloc[idx-count_num]['거래량']
            등락률 = data1.iloc[idx-count_num]['등락률']
            대비 = data1.iloc[idx-count_num]['대비']
            이평5 = data1.iloc[idx-count_num]['5이평']
            이평10 = data1.iloc[idx-count_num]['10이평']
            이평20 = data1.iloc[idx-count_num]['20이평']
            이평60 = data1.iloc[idx-count_num]['60이평']
            이평120 = data1.iloc[idx-count_num]['120이평']
            이평240 = data1.iloc[idx-count_num]['240이평']
            이격도20 = (종가 / 이평20) * 100
            이격도60 = (종가 / 이평60) * 100
            이격도120 = (종가 / 이평120) * 100
            try:
                금융투자 = data3.iloc[idx_3-count_num]['금융투자']
                보험 = data3.iloc[idx_3-count_num]['보험']
                투신 = data3.iloc[idx_3-count_num]['투신']
                사모 = data3.iloc[idx_3-count_num]['사모']
                은행 = data3.iloc[idx_3-count_num]['은행']
                기타금융 = data3.iloc[idx_3-count_num]['기타금융']
                연기금등 = data3.iloc[idx_3-count_num]['연기금 등']
                기타법인 = data3.iloc[idx_3-count_num]['기타법인']
                개인 = data3.iloc[idx_3-count_num]['개인']
                외국인 = data3.iloc[idx_3-count_num]['외국인']
                기타외국인 = data3.iloc[idx_3-count_num]['기타외국인']
            except:
                금융투자 = 'non'
                보험 = 'non'
                투신 = 'non'
                사모 = 'non'
                은행 = 'non'
                기타금융 = 'non'
                연기금등 = 'non'
                기타법인 = 'non'
                개인 = 'non'
                외국인 = 'non'
                기타외국인 = 'non'
                
            # 1씩 증가
            count_num -= 1
            
            ##---------------------------------------------------------------------##
            # 셀에 값입력
            3,139
            # 좌표로 값 출력 = sheet.cell(행, 열, 값)
            sheet.cell(115, 3+count_col, 지수시가)
            sheet.cell(116, 3+count_col, 지수고가)
            sheet.cell(117, 3+count_col, 지수저가)
            sheet.cell(118, 3+count_col, 지수종가)
            sheet.cell(119, 3+count_col, 시가)
            sheet.cell(120, 3+count_col, 고가)
            sheet.cell(121, 3+count_col, 저가)
            sheet.cell(122, 3+count_col, 종가)
            sheet.cell(123, 3+count_col, 거래량)
            sheet.cell(124, 3+count_col, 등락률)
            sheet.cell(125, 3+count_col, 대비) 
            sheet.cell(126, 3+count_col, 이평5)
            sheet.cell(127, 3+count_col, 이평10)
            sheet.cell(128, 3+count_col, 이평20)
            sheet.cell(129, 3+count_col, 이평60)
            sheet.cell(130, 3+count_col, 이평120)
            sheet.cell(131, 3+count_col, 이평240)
            sheet.cell(132, 3+count_col, 이격도20)
            sheet.cell(133, 3+count_col, 이격도60)
            sheet.cell(134, 3+count_col, 이격도120)
            sheet.cell(142, 3+count_col, 금융투자)
            sheet.cell(143, 3+count_col, 보험)
            sheet.cell(144, 3+count_col, 투신)
            sheet.cell(145, 3+count_col, 사모)
            sheet.cell(146, 3+count_col, 은행)
            sheet.cell(147, 3+count_col, 기타금융)
            sheet.cell(148, 3+count_col, 연기금등)
            sheet.cell(149, 3+count_col, 기타법인)
            sheet.cell(151, 3+count_col, 개인)
            sheet.cell(152, 3+count_col, 외국인)
            sheet.cell(153, 3+count_col, 기타외국인)        
            
            # 행 1씩 증가
            count_col += 1

        
        ##---------------------------------------------------------------------##
        # AA26~AD26에 고점, 저점 값 넣기
        arr_ago = [20, 60, 120, 240]
        
        # 행,열 이동
        col_num = 0
        
        for i in arr_ago :
            # 고가를 배열로 만들기
            arr_고가 = data1[idx-i:idx]['고가']
            # 고가에서 0은 제외
            arr_고가 = [x for x in arr_고가 if x != 0]
            # 주가최대구하기
            try:
                주가최대 = max(arr_고가)
            except:
                주가최대 = 'nondata'
            
            # 저가를 배열로 만들기
            arr_저가 = data1[idx-i:idx]['저가']
            # 저가에서 0은 제외
            arr_저가 = [x for x in arr_저가 if x != 0]
            # 주가최소구하기
            try:
                주가최소 = min(arr_저가)
            except:
                주가최소 = 'nondata'
            
            # 셀에 표시
            sheet.cell(64,27+col_num,주가최대)
            sheet.cell(66,27+col_num,주가최소)
           
            # 열이동
            col_num += 1
           
        ##---------------------------------------------------------------------##        
        # 저장
        wb.save('analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
        
#conn = chart_info_cls()
#conn.exe_chart_info('셀트리온',20200504)
