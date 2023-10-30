import openpyxl
import pandas as pd
from datetime import datetime
import mplfinance as mpf
from matplotlib import gridspec
from _ast import If
from matplotlib.pyplot import xlabel
from openpyxl.drawing.image import Image

class GBA2_cls:
    def Exe_GBA2(self,name,dd):
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('F:/JusikData/analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']

        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "F:/JusikData/oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        #print(data1.columns)
        
        
        
        ##---------------------------------------------------------------------##
        # 그래프 만들기
        # 필요한 열만 가져오기
        # 0(일자),8(시가),9(고가),10(저가),5(종가),11(거래량)
        df = data1.iloc[:,[0,8,9,10,5,11]]
        
        # 일자를 datetime형식으로 바꾸기
        x_data_date = df['일자'].values.tolist()
        
        # 일자를 datetime 형식으로 바꾼뒤 배열에 저장하기
        x_date_arr = []
        for i in x_data_date:
            # 일자만 배열에 저장
            sl_date = str(i)[0:4] + '-' + str(i)[4:6] + '-' + str(i)[6:8]
            real_date = datetime.strptime(sl_date, "%Y-%m-%d").date()
            x_date_arr.append(sl_date)
        
        # df의 인덱스를 datetime 형식으로 바꾸기
        df.index = pd.to_datetime(x_date_arr)
        
        # 일자 열은 없애기
        df = df.drop(['일자'],axis=1)
        
        # 컬럼명 바꾸기 - 바꿔야 그려짐
        df.columns = ['Open', 'High', 'Low', 'Close', 'Volume']
        
        # 차트 설정
        colorset = mpf.make_marketcolors(up='tab:red', down='tab:blue', volume='tab:blue')
        s = mpf.make_mpf_style(base_mpf_style='default', marketcolors=colorset)
        
        # 그래프의 시작과 끝 정하기
        today_index = data1.loc[data1['일자'] == dd]['일자'].index[0]
        arr_start = [today_index, 240, 120, 60, 20]
        
        grape_num = 1
        for i in arr_start:
            if i == today_index:
                fig_scl = 2.2
                fig_ra_s = 12
                fig_ra_e = 4
            else:
                fig_scl = 1.0
                fig_ra_s = 12
                fig_ra_e = 4
                
            # 그래프 그리기
            mpf.plot(df[today_index-i+1:today_index+1],              # 데이터 지정
                     volume=True,           # 거래량 표시 지정
                     mav=(20,60,120,240),   # 이동평균선 지정
                     figscale=fig_scl,          # 그림 크기 지정
                     figratio=(fig_ra_s,fig_ra_e),       # 그림 높이, 너비 비율 지정
                     savefig=dict(fname='F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_'+str(grape_num)+'_일봉차트.png',dpi=100,pad_inches=0.25),    # 저장
                     type='candle',         # 차트 종류 지정
                     style=s,               # 차트 스타일 지정(blueskies, checkers, classic, default, sas)
                     tight_layout=True,     # 레이아웃 지정
                     ylabel='Price')        # y축 라벨 지정
            
            #이미지 경로지정, 이미지 붙이기
            img_path = 'F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_'+str(grape_num)+'_일봉차트.png'
            img = Image(img_path)
            
            if grape_num == 1:
                sheet.add_image(img,'W48')
            elif grape_num == 2:
                sheet.add_image(img,'W83')
            elif grape_num == 3:
                sheet.add_image(img,'AJ83')
            elif grape_num == 4:
                sheet.add_image(img,'W105')
            elif grape_num == 5:
                sheet.add_image(img,'AJ105')
                
            grape_num += 1
        ##---------------------------------------------------------------------##
        
        
        
        ##---------------------------------------------------------------------##
        # 기간동안 주가, 거래량, 이격도 - 최대값, 최소값
        # 오늘의 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        arr_ago = [idx,720,480,240,120,60,20]
        
        # 이동평균선 만들기
        ma20 = data1['종가'].rolling(window=20).mean()
        ma60 = data1['종가'].rolling(window=60).mean()
        ma120 = data1['종가'].rolling(window=120).mean()
        ma240 = data1['종가'].rolling(window=240).mean()
        
        # data1에 이격도 데이터 합치기
        data1['종_20'] = (data1['종가'] / ma20) * 100
        data1['종_60'] = (data1['종가'] / ma60) * 100
        data1['20_120'] = (ma20 / ma120) * 100
        data1['20_240'] = (ma20 / ma240) * 10035
        
        print(data1)
        
        # 행,열 이동
        col_num = 0
        
        for i in arr_ago :
            # 고가를 배열로 만들기
            arr_고가 = data1[idx-i+1:idx+1]['고가']
            # 고가에서 0은 제외
            arr_고가 = [x for x in arr_고가 if x != 0]
            # 주가최대구하기
            try:
                주가최대 = max(arr_고가)
            except:
                주가최대 = 'nondata'
            
            # 저가를 배열로 만들기
            arr_저가 = data1[idx-i+1:idx+1]['저가']
            # 저가에서 0은 제외
            arr_저가 = [x for x in arr_저가 if x != 0]
            # 주가최소구하기
            try:
                주가최소 = min(arr_저가)
            except:
                주가최소 = 'nondata'
            
            # 거래량 배열로 만들기
            arr_거래량최대 = data1[idx-i+1:idx+1]['거래량'] * data1[idx-i+1:idx+1]['종가']
            # 0은 제외
            arr_거래량최대 = [x for x in arr_거래량최대 if x != 0]
            # 최대구하기
            try:
                거래량최대 = max(arr_거래량최대)
            except:
                거래량최대 = 'nondata'
        
            # 거래량을 배열로 만들기
            arr_거래량최소 = data1[idx-i+1:idx+1]['거래량'] * data1[idx-i+1:idx+1]['종가']
            # 0은 제외
            arr_거래량최소 = [x for x in arr_거래량최소 if x != 0]
            # 최소구하기
            try:
                거래량최소 = min(arr_거래량최소)
            except:
                거래량최소 = 'nondata'
            
            # 이격도 최대
            종_20최대 = data1[idx-i+1:idx+1]['종_20'].max()
            종_60최대 = data1[idx-i+1:idx+1]['종_60'].max()
            max_20_120 = data1[idx-i+1:idx+1]['20_120'].max()
            max_20_240 = data1[idx-i+1:idx+1]['20_240'].max()
            
            # 이격도 최소
            종_20최소 = data1[idx-i+1:idx+1]['종_20'].min()
            종_60최소 = data1[idx-i+1:idx+1]['종_60'].min() 
            min_20_120 = data1[idx-i+1:idx+1]['20_120'].min()
            min_20_240 = data1[idx-i+1:idx+1]['20_240'].min()
            ##-----------------------------------------##
            # 셀에 표시
            sheet.cell(132,24+col_num,주가최대)
            sheet.cell(134,24+col_num,주가최소)
            sheet.cell(142,24+col_num,종_20최대)
            sheet.cell(143,24+col_num,종_60최대)
            sheet.cell(144,24+col_num,max_20_120)
            sheet.cell(145,24+col_num,max_20_240)
            sheet.cell(146,24+col_num,거래량최대)
            sheet.cell(142,31+col_num,종_20최소)
            sheet.cell(143,31+col_num,종_60최소)
            sheet.cell(144,31+col_num,min_20_120)
            sheet.cell(145,31+col_num,min_20_240)
            sheet.cell(146,31+col_num,거래량최소)
            
            # 열 한칸씩 이동
            col_num += 1
                    
        # 당일 이격도
        dday_종_20 = data1.iloc[idx]['종_20']
        dday_종_60 = data1.iloc[idx]['종_60']
        dday_20_120 = data1.iloc[idx]['20_120']
        dday_20_240 = data1.iloc[idx]['20_240']
        ##-----------------------------------##
        # 셀에 표시
        sheet.cell(128,32,dday_종_20)
        sheet.cell(128,34,dday_종_60)        
        sheet.cell(128,36,dday_20_120)
        sheet.cell(128,38,dday_20_240)



        ##---------------------------------------------------------------------##
        # 저장
        wb.save('F:/JusikData/analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
#conn = GBA2_cls()
#conn.Exe_GBA2('삼성전자',20200504)