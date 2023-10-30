import openpyxl
import pandas as pd
import os
import datetime
import sys

class GBN1_cls:
    # name = 종목명, dd = 일자, q = 분기
    def Exe_GBN1(self,name,dd,q):
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('F:/JusikData/analysis_csv/HJS/개별분석양식.xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']


        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "F:/JusikData/oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        #print(data1.columns)
        
        ## data4 = 업종 분류 데이터 ##
        path4 = "F:/JusikData/oneday_csv/data_sectors/"+str(dd)[0:4]+'.csv'
        data4 = pd.read_csv(path4, encoding='cp949')
        #print(data4.columns)
        
        
        
        ##---------------------------------------------------------------------##
        ## 변수에 데이터 저장
        # 해당일의 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        
        # 변수에 데이터 저장
        매수일자 = dd
        소속부 = data1.iloc[idx]['소속부']
        종목명 = name
        분기 = q # 분기는 직접 입력(dart에 자료 있는지 여부)
        상장주식수 = data1.iloc[idx]['상장주식수']
        try:
            업종구분 = data4.loc[data4['종목명'] == name]['업종명'].values[0]
        except:
            업종구분 = 'non data'
        상장시장 = data1.iloc[idx]['시장구분']
        시가총액 = data1.iloc[idx]['시가총액']
        시가 = data1.iloc[idx]['시가']
        고가 = data1.iloc[idx]['고가']
        저가 = data1.iloc[idx]['저가']
        종가 = data1.iloc[idx]['종가']
        거래량 = data1.iloc[idx]['거래량']
        거래대금 = data1.iloc[idx]['거래대금']
        
        
        
        ##---------------------------------------------------------------------##
        # 좌표로 값 출력 = sheet.cell(행, 열, 값)
        sheet.cell(20, 3, 매수일자)
        sheet.cell(20, 5, 분기)
        sheet.cell(20, 7, 소속부)
        sheet.cell(20, 11, 상장주식수)
        sheet.cell(21, 3, 종목명)
        sheet.cell(21, 7, 업종구분)
        sheet.cell(21, 11, 상장시장)
        sheet.cell(21, 13, 시가총액)
        sheet.cell(22, 3, 시가)
        sheet.cell(22, 5, 고가)
        sheet.cell(22, 7, 저가)
        sheet.cell(22, 9, 종가)
        sheet.cell(22, 11, 거래량)
        sheet.cell(22, 14, 거래대금)
        ##---------------------------------------------------------------------##
        
        
        # 저장
        wb.save('F:/JusikData/analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
#conn = GBN1_cls()
#conn.Exe_GBN1('삼성전자',20200504,1)