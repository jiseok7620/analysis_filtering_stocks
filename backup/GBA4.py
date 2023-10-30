import openpyxl
import pandas as pd
from datetime import datetime
import numpy as np
import mplfinance as mpf
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

class GBA4_cls:
    def make_obv(self, name, data1, dd, display):
        OBV = []
        OBV.append(0)
        
        for i in data1.index:
            if data1.iloc[i]['종가'] > data1.iloc[i-1]['종가'] :
                OBV.append(OBV[-1] + data1.iloc[i]['거래량'])
            elif data1.iloc[i]['종가'] < data1.iloc[i-1]['종가'] :
                OBV.append(OBV[-1] - data1.iloc[i]['거래량'])
            else:
                OBV.append(OBV[-1])
        
        # OBV에서 처음에 넣어줬던 0을 삭제하기
        del OBV[OBV.index(0)]
        
        # 현재날짜의 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        
        # 캔버스 설정(크기, 배경 설정 등)
        fig = plt.figure(figsize=(15,5)) ## 캔버스 생성
        fig.set_facecolor('white') ## 캔버스 색상 설정
        ax = fig.add_subplot() ## 그림 뼈대(프레임) 생성
        
        # 선그래프 그리기
        ax.plot(OBV[idx-display-1:idx+1], color='red', marker='o', markersize=3)
        
        # 그래프 보기
        #plt.show()
        
        # 저장
        plt.savefig('F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_obv차트.png')
    
        # 메모리 제거
        plt.close()
        
    def make_vr(self, name, data1, dd, display, period):
        # 현재날짜의 인덱스
        idx = data1.loc[data1['일자'] == dd]['일자'].index[0]
        
        # VR 배열
        VR = []
        
        for i in data1.index:
            if i >= period :
                # 상승분, 하락분 배열
                VR_up = []
                VR_down = []
                VR_up.append(0)
                VR_down.append(0)
                
                for j in range(i-period+1,i+1):
                    if data1.iloc[j]['종가'] > data1.iloc[j-1]['종가'] :
                        VR_up.append(VR_up[-1] + data1.iloc[j]['거래량'])
                        VR_down.append(VR_down[-1])
                    elif data1.iloc[j]['종가'] < data1.iloc[j-1]['종가'] :
                        VR_up.append(VR_up[-1])
                        VR_down.append(VR_down[-1] + data1.iloc[j]['거래량'])
                    else:
                        VR_up.append(VR_up[-1] + data1.iloc[j]['거래량']/2)
                        VR_down.append(VR_down[-1] + data1.iloc[j]['거래량']/2)
                
                VR.append((VR_up[-1] / VR_down[-1]) *100)
                
            else :
                VR.append(0)
        
        # 캔버스 설정(크기, 배경 설정 등)
        fig = plt.figure(figsize=(15,5)) ## 캔버스 생성
        fig.set_facecolor('white') ## 캔버스 색상 설정
        ax = fig.add_subplot() ## 그림 뼈대(프레임) 생성
        
        # 선그래프 그리기
        ax.plot(VR[idx-display-1:idx+1], color='red', marker='o', markersize=3)
        
        # 0에 수평선 긋기
        ax.axhline(150,color='green') 
        ax.axhline(100,color='green')
        
        # 그래프 보기
        #plt.show()
        
        # 저장
        plt.savefig('F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_vr차트.png')
    
        # 메모리 제거
        plt.close()
        
    def Exe_GBA4(self, name, dd):
        # 엑셀 파일 열기
        wb = openpyxl.load_workbook('F:/JusikData/analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        
        # 시트 지정 하기
        sheet = wb['Sheet1']


        ##---------------------------------------------------------------------##
        # 데이터 가져오기
        ## data1 = 개별 종목 데이터 ##
        path1 = "F:/JusikData/oneday_csv/onedaydata/"+name+'/'+name+'.csv'
        data1 = pd.read_csv(path1, encoding='cp949')
        print(data1.columns)
        
        ## data3 = 수급 데이터 ##
        path3 = "F:/JusikData/oneday_csv/onedaydata_supply/"+name+'.csv'
        data3 = pd.read_csv(path3, encoding='cp949')
        print(data3.columns)
        
        
        
        ##---------------------------------------------------------------------##
        conn = GBA4_cls()
        
        # OBV 그래프 만들기
        conn.make_obv(name, data1, dd, 60)
        
        # VR 그래프 만들기
        conn.make_vr(name, data1, dd, 100, 60)
        
        # 그래프 엑셀에 붙여넣기
        #이미지 경로지정, 이미지 붙이기
        # 1. roc 차트
        img_path1 = 'F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_obv차트.png'
        img1 = Image(img_path1)
            
        sheet.add_image(img1,'W189')
            
        # 1. rsi 차트
        img_path2 = 'F:/JusikData/analysis_csv/HJS/img/'+name+'_'+str(dd)+'_vr차트.png'
        img2 = Image(img_path2)
            
        sheet.add_image(img2,'AJ189')
        
        
        
        ##---------------------------------------------------------------------##
        # 변수에 데이터 저장
        
        # data3의 일자 형식 바꾸기
        dd_y = int(str(dd)[0:4])
        dd_m = int(str(dd)[4:6])
        dd_d = int(str(dd)[6:8])
        if len(str(dd_m)) == 1:
            if len(str(dd_d)) == 1:
                dd_full = str(dd_y)[0:4] + "/0" + str(dd_m) + "/0" + str(dd_d)
            else:
                dd_full = str(dd_y)[0:4] + "/0" + str(dd_m) + "/" + str(dd_d)
        else :
            if len(str(dd_d)) == 1:
                dd_full = str(dd_y)[0:4] + "/" + str(dd_m) + "/0" + str(dd_d)
            else:
                dd_full = str(dd_y)[0:4] + "/" + str(dd_m) + "/" + str(dd_d)
            
        # 현재 날짜의 인덱스
        idx = data3.loc[data3['일자'] == dd_full]['일자'].index[0]
        
        # 행, 열 이동
        col_num = 0
        row_num = 0
        
        for i in range(idx-60,idx+1):
            금융투자 = data3.iloc[i]['금융투자']
            보험 = data3.iloc[i]['보험']
            투신 = data3.iloc[i]['투신']
            사모 = data3.iloc[i]['사모']
            은행 = data3.iloc[i]['은행']
            기타금융 = data3.iloc[i]['기타금융']
            연기금등 = data3.iloc[i]['연기금 등']
            기타법인 = data3.iloc[i]['기타법인']
            개인 = data3.iloc[i]['개인']
            외국인 = data3.iloc[i]['외국인']
            기타외국인 = data3.iloc[i]['기타외국인']

            ##---------------------------------------------------------------------##
            # 좌표로 값 출력 = sheet.cell(행, 열, 값)
            sheet.cell(279-row_num, 24, 금융투자)
            sheet.cell(279-row_num, 25, 보험)
            sheet.cell(279-row_num, 26, 투신)
            sheet.cell(279-row_num, 27, 사모)
            sheet.cell(279-row_num, 28, 은행)
            sheet.cell(279-row_num, 29, 기타금융)
            sheet.cell(279-row_num, 30, 연기금등)
            sheet.cell(279-row_num, 31, 기타법인)
            sheet.cell(279-row_num, 33, 개인)
            sheet.cell(279-row_num, 34, 외국인)
            sheet.cell(279-row_num, 35, 기타외국인)
            
            
            # 행 열 이동
            row_num += 1
            
            
        
        
        
        ##---------------------------------------------------------------------##
        # 저장
        wb.save('F:/JusikData/analysis_csv/HJS/'+name+'_'+str(dd)+'.xlsx')
        wb.close()
        
#conn = GBA4_cls()
#conn.Exe_GBA4('삼성전자',20200504)